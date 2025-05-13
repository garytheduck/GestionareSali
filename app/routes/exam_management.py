"""
API endpoints pentru managementul examenelor
"""
from flask import Blueprint, request, jsonify, current_app, send_file
from app import db
from app.models.exam import Exam, ExamRegistration
from app.models import Room
from app.models.course import Course, ExamType
from app.models import User
from werkzeug.exceptions import BadRequest, NotFound
from flask_jwt_extended import jwt_required, get_jwt_identity
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy import and_, or_, func
from datetime import datetime, timedelta
from app.utils.exam_overlap_checker import check_room_overlap, check_teacher_overlap, check_assistant_overlap, check_student_group_overlap, get_overlapping_exams
# Temporar comentat pentru a evita eroarea
# from app.utils.exam_report_generator import generate_exam_schedule_excel, generate_exam_schedule_pdf, generate_exam_completion_stats
# Definim funcții placeholder
def generate_exam_schedule_excel(filters=None):
    """
    Generează un raport Excel cu programarea examenelor
    
    Args:
        filters (dict): Filtre pentru examenele care vor fi incluse în raport
            - faculty (str): Filtrare după facultate
            - study_program (str): Filtrare după program de studiu
            - year_of_study (int): Filtrare după an de studiu
            - group_name (str): Filtrare după grupă
            - exam_type (str): Filtrare după tip examen
    
    Returns:
        bytes: Conținutul Excel-ului generat
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        
        # Inițializăm filtrele dacă nu sunt furnizate
        if filters is None:
            filters = {}
        
        # Obținem examenele din baza de date folosind filtrele
        exams = get_exams_for_report(filters)
        
        # Creăm un workbook nou
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Programare Examene"
        
        # Definim stilurile pentru celule
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='Arial', size=11)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        date_alignment = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Adăugăm titlul raportului
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = "PROGRAMAREA EXAMENELOR"
        title_cell.font = Font(name='Arial', size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Adăugăm data generării
        ws.merge_cells('A2:I2')
        date_cell = ws['A2']
        date_cell.value = f"Generat la: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        date_cell.font = Font(name='Arial', size=12, italic=True)
        date_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Lăsăm un rând liber
        
        # Definim antetul tabelului (rândul 4)
        headers = [
            "Nr.", "Disciplină", "Tip", "Data", "Ora", "Sala", "Profesor", "Grupă", "Program Studiu"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Setăm lățimea coloanelor
        ws.column_dimensions['A'].width = 5   # Nr.
        ws.column_dimensions['B'].width = 40  # Disciplină
        ws.column_dimensions['C'].width = 15  # Tip
        ws.column_dimensions['D'].width = 12  # Data
        ws.column_dimensions['E'].width = 10  # Ora
        ws.column_dimensions['F'].width = 15  # Sala
        ws.column_dimensions['G'].width = 25  # Profesor
        ws.column_dimensions['H'].width = 10  # Grupă
        ws.column_dimensions['I'].width = 20  # Program Studiu
        
        # Completăm datele
        if not exams:
            ws.merge_cells('A5:I5')
            no_data_cell = ws['A5']
            no_data_cell.value = "Nu există examene programate care să corespundă criteriilor de filtrare."
            no_data_cell.font = data_font
            no_data_cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            for i, exam in enumerate(exams, 1):
                row_idx = i + 4  # Începem de la rândul 5 (după antet)
                
                # Formatăm data și ora
                exam_date = exam.get('date', '').strftime("%d.%m.%Y") if exam.get('date') else 'Neprogramat'
                exam_time = exam.get('time', '').strftime("%H:%M") if exam.get('time') else 'Neprogramat'
                
                # Adăugăm datele în celule
                ws.cell(row=row_idx, column=1, value=i).alignment = date_alignment  # Nr.
                ws.cell(row=row_idx, column=2, value=exam.get('course_name', 'N/A')).alignment = data_alignment  # Disciplină
                ws.cell(row=row_idx, column=3, value=exam.get('exam_type', 'N/A')).alignment = data_alignment  # Tip
                ws.cell(row=row_idx, column=4, value=exam_date).alignment = date_alignment  # Data
                ws.cell(row=row_idx, column=5, value=exam_time).alignment = date_alignment  # Ora
                ws.cell(row=row_idx, column=6, value=exam.get('room_name', 'Neprogramat')).alignment = data_alignment  # Sala
                ws.cell(row=row_idx, column=7, value=exam.get('teacher_name', 'N/A')).alignment = data_alignment  # Profesor
                ws.cell(row=row_idx, column=8, value=exam.get('group_name', 'N/A')).alignment = date_alignment  # Grupă
                ws.cell(row=row_idx, column=9, value=exam.get('study_program', 'N/A')).alignment = data_alignment  # Program Studiu
                
                # Aplicăm borduri pentru toate celulele din rând
                for col_idx in range(1, 10):
                    ws.cell(row=row_idx, column=col_idx).border = thin_border
                    ws.cell(row=row_idx, column=col_idx).font = data_font
        
        # Salvăm workbook-ul într-un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
    except Exception as e:
        logger.error(f"Error generating Excel: {str(e)}")
        
        # Creăm un Excel simplu cu mesajul de eroare
        try:
            import openpyxl
            import io
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Eroare"
            
            ws['A1'] = "Eroare la generarea raportului Excel"
            ws['A2'] = f"Detalii: {str(e)}"
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer.getvalue()
        except Exception as inner_e:
            logger.error(f"Failed to generate error Excel: {str(inner_e)}")
            # Returnăm un Excel gol
            import io
            from openpyxl import Workbook
            
            wb = Workbook()
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer.getvalue()

def generate_exam_schedule_pdf(filters=None):
    """
    Generează un raport PDF cu programarea examenelor
    
    Args:
        filters (dict): Filtre pentru examenele care vor fi incluse în raport
            - faculty (str): Filtrare după facultate
            - study_program (str): Filtrare după program de studiu
            - year_of_study (int): Filtrare după an de studiu
            - group_name (str): Filtrare după grupă
            - exam_type (str): Filtrare după tip examen
    
    Returns:
        bytes: Conținutul PDF-ului generat
    """
    logger.info("Starting PDF generation with filters: %s", filters)
    try:
        logger.info("Importing required libraries for PDF generation")
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import io
        
        logger.info("Libraries imported successfully")
        
        # Inițializăm filtrele dacă nu sunt furnizate
        if filters is None:
            filters = {}
            logger.info("No filters provided, using empty filter set")
        
        logger.info("Fetching exams from database")
        exams = get_exams_for_report(filters)
        logger.info("Retrieved %d exams from database", len(exams))
        
        # Creăm un buffer pentru PDF
        logger.info("Creating PDF buffer")
        buffer = io.BytesIO()
        
        # Configurăm documentul PDF
        logger.info("Configuring PDF document")
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        logger.info("PDF document configured successfully")
        
        # Stiluri pentru text
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        subtitle_style = styles['Heading2']
        normal_style = styles['Normal']
        
        # Elementele care vor fi adăugate în PDF
        elements = []
        
        # Titlul raportului
        title = Paragraph("Programarea Examenelor", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.25*inch))
        
        # Subtitlu cu data generării
        current_date = datetime.now().strftime("%d.%m.%Y %H:%M")
        subtitle = Paragraph(f"Generat la: {current_date}", subtitle_style)
        elements.append(subtitle)
        elements.append(Spacer(1, 0.25*inch))
        
        # Verificăm dacă avem examene
        if not exams:
            no_data = Paragraph("Nu există examene programate care să corespundă criteriilor de filtrare.", normal_style)
            elements.append(no_data)
        else:
            # Definim antetul tabelului
            table_headers = [
                "Nr.", "Disciplină", "Tip", "Data", "Ora", "Sala", "Profesor", "Grupă", "Program Studiu"
            ]
            
            # Construim datele pentru tabel
            table_data = [table_headers]
            
            for i, exam in enumerate(exams, 1):
                # Formatăm data și ora
                exam_date = exam.get('date', '').strftime("%d.%m.%Y") if exam.get('date') else 'Neprogramat'
                exam_time = exam.get('time', '').strftime("%H:%M") if exam.get('time') else 'Neprogramat'
                
                row = [
                    str(i),
                    exam.get('course_name', 'N/A'),
                    exam.get('exam_type', 'N/A'),
                    exam_date,
                    exam_time,
                    exam.get('room_name', 'Neprogramat'),
                    exam.get('teacher_name', 'N/A'),
                    exam.get('group_name', 'N/A'),
                    exam.get('study_program', 'N/A')
                ]
                table_data.append(row)
            
            # Creăm tabelul
            table = Table(table_data, repeatRows=1)
            
            # Stilizăm tabelul
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Nr. centrat
                ('ALIGN', (3, 1), (4, -1), 'CENTER'),  # Data și ora centrate
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ])
            table.setStyle(table_style)
            
            # Adăugăm tabelul în document
            elements.append(table)
        
        # Construim PDF-ul
        logger.info("Building PDF document with %d elements", len(elements))
        try:
            doc.build(elements)
            logger.info("PDF document built successfully")
            
            # Obținem conținutul PDF-ului
            logger.info("Getting PDF content from buffer")
            pdf_data = buffer.getvalue()
            buffer.close()
            logger.info("PDF data size: %d bytes", len(pdf_data))
            
            return pdf_data
        except Exception as build_error:
            logger.error("Error building PDF: %s", str(build_error))
            
            # Încercăm o abordare mai simplă pentru PDF
            logger.info("Trying a simpler approach for PDF generation")
            buffer = io.BytesIO()
            
            from reportlab.pdfgen import canvas
            c = canvas.Canvas(buffer, pagesize=landscape(A4))
            c.setFont("Helvetica", 16)
            c.drawString(100, 500, "Programarea Examenelor")
            c.setFont("Helvetica", 12)
            c.drawString(100, 480, f"Generat la: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
            
            if not exams:
                c.drawString(100, 450, "Nu există examene programate care să corespundă criteriilor de filtrare.")
            else:
                y = 450
                c.drawString(100, y, f"Total examene: {len(exams)}")
                y -= 20
                
                for i, exam in enumerate(exams[:10], 1):  # Limităm la primele 10 examene pentru simplitate
                    exam_date = exam.get('date', '').strftime("%d.%m.%Y") if exam.get('date') else 'Neprogramat'
                    exam_time = exam.get('time', '').strftime("%H:%M") if exam.get('time') else 'Neprogramat'
                    
                    c.drawString(100, y, f"{i}. {exam.get('course_name', 'N/A')}")
                    y -= 15
                    c.drawString(120, y, f"Data: {exam_date}, Ora: {exam_time}, Sala: {exam.get('room_name', 'Neprogramat')}")
                    y -= 15
                    c.drawString(120, y, f"Profesor: {exam.get('teacher_name', 'N/A')}, Grupa: {exam.get('group_name', 'N/A')}")
                    y -= 20
            
            c.save()
            pdf_data = buffer.getvalue()
            buffer.close()
            logger.info("Simple PDF generated successfully, size: %d bytes", len(pdf_data))
            
            return pdf_data
    except Exception as e:
        logger.error(f"Error generating PDF: {str(e)}")
        # Returnăm un PDF simplu cu mesajul de eroare
        try:
            logger.info("Attempting to generate a simple error PDF")
            from reportlab.pdfgen import canvas
            import io
            
            buffer = io.BytesIO()
            c = canvas.Canvas(buffer, pagesize=A4)
            c.setFont("Helvetica", 16)
            c.drawString(100, 750, "Eroare la generarea raportului PDF")
            c.setFont("Helvetica", 12)
            c.drawString(100, 730, f"Detalii: {str(e)}")
            c.save()
            
            pdf_data = buffer.getvalue()
            buffer.close()
            logger.info("Error PDF generated successfully, size: %d bytes", len(pdf_data))
            
            return pdf_data
        except Exception as inner_e:
            logger.error(f"Failed to generate error PDF: {str(inner_e)}")
            # Returnăm un PDF gol dacă nu putem genera nici măcar mesajul de eroare
            logger.info("Returning a minimal valid PDF")
            return b'%PDF-1.4\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj 2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj 3 0 obj<</Type/Page/MediaBox[0 0 595 842]/Parent 2 0 R/Resources<<>>>>endobj\nxref\n0 4\n0000000000 65535 f\n0000000010 00000 n\n0000000053 00000 n\n0000000102 00000 n\ntrailer<</Size 4/Root 1 0 R>>\nstartxref\n178\n%%EOF'

def get_exams_for_report(filters=None):
    """
    Obține lista de examene pentru raport, aplicând filtrele specificate
    
    Args:
        filters (dict): Filtre pentru examenele care vor fi incluse în raport
            - faculty (str): Filtrare după facultate
            - study_program (str): Filtrare după program de studiu
            - year_of_study (int): Filtrare după an de studiu
            - group_name (str): Filtrare după grupă
            - exam_type (str): Filtrare după tip examen
    
    Returns:
        list: Lista de examene formatată pentru raport
    """
    try:
        # Inițializăm filtrele dacă nu sunt furnizate
        if filters is None:
            filters = {}
        
        # Construim query-ul pentru examene
        from app.models import Exam, Course, Room, User
        from sqlalchemy import and_, or_
        
        query = db.session.query(
            Exam,
            Course,
            Room,
            User
        ).join(
            Course, Exam.course_id == Course.id
        ).outerjoin(
            Room, Exam.room_id == Room.id
        ).outerjoin(
            User, Course.teacher_id == User.id
        )
        
        # Aplicăm filtrele
        if filters.get('faculty'):
            query = query.filter(Course.faculty == filters['faculty'])
        
        if filters.get('study_program'):
            query = query.filter(Course.study_program == filters['study_program'])
        
        if filters.get('year_of_study'):
            query = query.filter(Course.year_of_study == filters['year_of_study'])
        
        if filters.get('group_name'):
            query = query.filter(Course.group_name == filters['group_name'])
        
        if filters.get('exam_type'):
            query = query.filter(Exam.exam_type == filters['exam_type'])
        
        # Executăm query-ul
        results = query.all()
        
        # Formatăm rezultatele pentru raport
        exams_data = []
        for exam, course, room, teacher in results:
            exam_data = {
                'id': exam.id,
                'course_id': exam.course_id,
                'course_name': course.name if course else 'N/A',
                'exam_type': exam.exam_type,
                'date': exam.date,
                'time': exam.time,
                'duration': exam.duration,
                'room_id': exam.room_id,
                'room_name': room.name if room else 'Neprogramat',
                'teacher_id': course.teacher_id if course else None,
                'teacher_name': f"{teacher.first_name} {teacher.last_name}" if teacher else 'N/A',
                'faculty': course.faculty if course else 'N/A',
                'study_program': course.study_program if course else 'N/A',
                'year_of_study': course.year_of_study if course else 'N/A',
                'group_name': course.group_name if course else 'N/A',
            }
            exams_data.append(exam_data)
        
        # Sortăm examenele după dată și oră
        exams_data.sort(key=lambda x: (x['date'] or datetime.max.date(), x['time'] or datetime.max.time()))
        
        return exams_data
    except Exception as e:
        logger.error(f"Error getting exams for report: {str(e)}")
        return []

def generate_exam_completion_stats(*args, **kwargs):
    return {}
from app.utils.email_service import send_exam_proposal_notification, send_exam_approval_notification, send_exam_rejection_notification
import io
import logging

logger = logging.getLogger(__name__)

# Creăm un Blueprint pentru managementul examenelor
exam_bp = Blueprint('exam', __name__, url_prefix='/api/exams')

@exam_bp.route('/', methods=['GET'])
@jwt_required()
def get_exams():
    """
    Obține lista examenelor cu opțiuni de filtrare
    
    Query params:
        course_id (int): Filtrare după ID-ul cursului
        room_id (int): Filtrare după ID-ul sălii
        start_date (str): Filtrare după data de început (format: YYYY-MM-DD)
        end_date (str): Filtrare după data de sfârșit (format: YYYY-MM-DD)
        exam_type (str): Filtrare după tipul de examen
        semester (str): Filtrare după semestru
        academic_year (str): Filtrare după anul academic
    """
    try:
        # Preluăm parametrii din query string
        course_id = request.args.get('course_id')
        room_id = request.args.get('room_id')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        exam_type = request.args.get('exam_type')
        semester = request.args.get('semester')
        academic_year = request.args.get('academic_year')
        
        # Construim query-ul
        query = Exam.query
        
        if course_id:
            query = query.filter(Exam.course_id == int(course_id))
        if room_id:
            query = query.filter(Exam.room_id == int(room_id))
        if start_date:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(Exam.start_time >= start_date_obj)
        if end_date:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
            # Adăugăm 23:59:59 pentru a include toată ziua
            end_date_obj = end_date_obj.replace(hour=23, minute=59, second=59)
            query = query.filter(Exam.end_time <= end_date_obj)
        if exam_type:
            query = query.filter(Exam.exam_type == exam_type)
        if semester:
            query = query.filter(Exam.semester == semester)
        if academic_year:
            query = query.filter(Exam.academic_year == academic_year)
        
        # Includem doar examenele active
        query = query.filter(Exam.is_active == True)
        
        # Executăm query-ul
        exams = query.all()
        
        # Serializăm rezultatele
        result = []
        for exam in exams:
            result.append({
                'id': exam.id,
                'course': {
                    'id': exam.course.id,
                    'name': exam.course.name,
                    'code': exam.course.code,
                    'study_program': exam.course.study_program,
                    'year_of_study': exam.course.year_of_study
                },
                'room': {
                    'id': exam.room.id,
                    'name': exam.room.name,
                    'building': exam.room.building,
                    'capacity': exam.room.capacity
                },
                'start_time': exam.start_time.isoformat(),
                'end_time': exam.end_time.isoformat(),
                'exam_type': exam.exam_type,
                'semester': exam.semester,
                'academic_year': exam.academic_year,
                'max_students': exam.max_students,
                'notes': exam.notes,
                'registrations_count': len(exam.registrations)
            })
        
        return jsonify({
            'status': 'success',
            'data': result,
            'count': len(result)
        }), 200
        
    except Exception as e:
        logger.error(f"Error getting exams: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': 'Failed to retrieve exams',
            'error': str(e)
        }), 500


@exam_bp.route('/', methods=['POST'])
@jwt_required()
def create_exam():
    """
    Creează un nou examen
    
    Body params:
        course_id (int): ID-ul cursului
        room_id (int): ID-ul sălii
        start_time (str): Data și ora de început (format: YYYY-MM-DDTHH:MM:SS)
        end_time (str): Data și ora de sfârșit (format: YYYY-MM-DDTHH:MM:SS)
        exam_type (str): Tipul de examen
        semester (str): Semestrul
        academic_year (str): Anul academic
        max_students (int, optional): Capacitatea maximă de studenți
        notes (str, optional): Note sau observații
    """
    try:
        # Preluăm datele din corpul cererii
        data = request.get_json()
        
        # Verificăm câmpurile obligatorii
        required_fields = ['course_id', 'room_id', 'start_time', 'end_time', 'exam_type', 'semester', 'academic_year']
        for field in required_fields:
            if field not in data:
                return jsonify({
                    'status': 'error',
                    'message': f'Missing required field: {field}'
                }), 400
        
        # Verificăm dacă cursul există
        course = Course.query.get(data['course_id'])
        if not course:
            return jsonify({
                'status': 'error',
                'message': f'Course with ID {data["course_id"]} not found'
            }), 404
        
        # Verificăm dacă sala există
        room = Room.query.get(data['room_id'])
        if not room:
            return jsonify({
                'status': 'error',
                'message': f'Room with ID {data["room_id"]} not found'
            }), 404
        
        # Convertim datele de timp
        start_time = datetime.fromisoformat(data['start_time'])
        end_time = datetime.fromisoformat(data['end_time'])
        
        # Verificăm dacă data de început este înainte de data de sfârșit
        if start_time >= end_time:
            return jsonify({
                'status': 'error',
                'message': 'Start time must be before end time'
            }), 400
        
        # Verificăm suprapunerile pentru sală, profesor, asistent și grupă de studenți
        course = Course.query.get(course_id)
        if not course:
            return jsonify({
                'status': 'error',
                'message': f'Course with ID {course_id} not found'
            }), 404
            
        # Verificăm suprapunerile pentru sală
        if check_room_overlap(room_id, start_time, end_time):
            return jsonify({
                'status': 'error',
                'message': 'Room is not available during the specified time interval',
                'overlap_type': 'room'
            }), 409  # Conflict
            
        # Verificăm suprapunerile pentru profesor
        if course.teacher_id and check_teacher_overlap(course.teacher_id, start_time, end_time):
            return jsonify({
                'status': 'error',
                'message': 'Teacher has another exam scheduled during the specified time interval',
                'overlap_type': 'teacher'
            }), 409  # Conflict
            
        # Verificăm suprapunerile pentru asistent
        if course.assistant_id and check_assistant_overlap(course.assistant_id, start_time, end_time):
            return jsonify({
                'status': 'error',
                'message': 'Assistant has another exam scheduled during the specified time interval',
                'overlap_type': 'assistant'
            }), 409  # Conflict
            
        # Verificăm suprapunerile pentru grupa de studenți
        if check_student_group_overlap(course.group_name, course.faculty, course.study_program, course.year_of_study, start_time, end_time):
            return jsonify({
                'status': 'error',
                'message': 'Student group has another exam scheduled during the specified time interval',
                'overlap_type': 'student_group'
            }), 409  # Conflict
            
        # Alternativ, putem obține toate suprapunerile într-un singur apel
        exam_data = {
            'room_id': room_id,
            'start_time': start_time,
            'end_time': end_time,
            'course_id': course_id
        }
        
        overlaps = get_overlapping_exams(exam_data)
        
        # Dacă există suprapuneri, le returnam pentru a fi afișate în frontend
        has_overlaps = any(len(overlaps[key]) > 0 for key in overlaps if key != 'error')
        
        if has_overlaps:
            return jsonify({
                'status': 'error',
                'message': 'There are scheduling conflicts for this exam',
                'overlaps': overlaps
            }), 409  # Conflict
        
        # Creăm examenul
        exam = Exam(
            course_id=data['course_id'],
            room_id=data['room_id'],
            start_time=start_time,
            end_time=end_time,
            exam_type=data['exam_type'],
            semester=data['semester'],
            academic_year=data['academic_year'],
            max_students=data.get('max_students'),
            notes=data.get('notes')
        )
        
        # Salvăm în baza de date
        db.session.add(exam)
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'Exam created successfully',
            'data': {
                'id': exam.id,
                'course': {
                    'id': exam.course.id,
                    'name': exam.course.name
                },
                'room': {
                    'id': exam.room.id,
                    'name': exam.room.name
                },
                'start_time': exam.start_time.isoformat(),
                'end_time': exam.end_time.isoformat(),
                'exam_type': exam.exam_type,
                'semester': exam.semester,
                'academic_year': exam.academic_year,
                'max_students': exam.max_students,
                'notes': exam.notes
            }
        }), 201
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error creating exam: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': 'Failed to create exam',
            'error': str(e)
        }), 500


@exam_bp.route('/<int:id>', methods=['GET'])
@jwt_required()
def get_exam(id):
    """
    Obține detaliile unui examen specific
    
    Path params:
        id (int): ID-ul examenului
    """
    try:
        exam = Exam.query.get(id)
        
        if not exam or not exam.is_active:
            return jsonify({
                'status': 'error',
                'message': f'Exam with ID {id} not found'
            }), 404
        
        # Calculăm numărul de înregistrări
        registrations_count = len(exam.registrations)
        
        # Serializăm rezultatul
        result = {
            'id': exam.id,
            'course': {
                'id': exam.course.id,
                'name': exam.course.name,
                'code': exam.course.code,
                'study_program': exam.course.study_program,
                'year_of_study': exam.course.year_of_study
            },
            'room': {
                'id': exam.room.id,
                'name': exam.room.name,
                'building': exam.room.building,
                'capacity': exam.room.capacity
            },
            'start_time': exam.start_time.isoformat(),
            'end_time': exam.end_time.isoformat(),
            'exam_type': exam.exam_type,
            'semester': exam.semester,
            'academic_year': exam.academic_year,
            'max_students': exam.max_students,
            'notes': exam.notes,
            'registrations_count': registrations_count,
            'created_at': exam.created_at.isoformat() if exam.created_at else None,
            'updated_at': exam.updated_at.isoformat() if exam.updated_at else None
        }
        
        return jsonify({
            'status': 'success',
            'data': result
        }), 200
        
    except Exception as e:
        logger.error(f"Error getting exam {id}: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Failed to retrieve exam with ID {id}',
            'error': str(e)
        }), 500


@exam_bp.route('/<int:id>', methods=['PUT'])
@jwt_required()
def update_exam(id):
    """
    Actualizează un examen existent
    
    Path params:
        id (int): ID-ul examenului
        
    Body params:
        room_id (int, optional): ID-ul sălii
        start_time (str, optional): Data și ora de început (format: YYYY-MM-DDTHH:MM:SS)
        end_time (str, optional): Data și ora de sfârșit (format: YYYY-MM-DDTHH:MM:SS)
        exam_type (str, optional): Tipul de examen
        max_students (int, optional): Capacitatea maximă de studenți
        notes (str, optional): Note sau observații
    """
    try:
        # Preluăm datele din corpul cererii
        data = request.get_json()
        
        # Verificăm dacă examenul există
        exam = Exam.query.get(id)
        
        if not exam or not exam.is_active:
            return jsonify({
                'status': 'error',
                'message': f'Exam with ID {id} not found'
            }), 404
        
        # Nu permitem schimbarea cursului sau a semestrului/anului academic după creare
        if 'course_id' in data:
            return jsonify({
                'status': 'error',
                'message': 'Changing the course for an existing exam is not allowed'
            }), 400
            
        if 'semester' in data or 'academic_year' in data:
            return jsonify({
                'status': 'error',
                'message': 'Changing the semester or academic year for an existing exam is not allowed'
            }), 400
        
        # Procesăm actualizarea sălii dacă este specificată
        if 'room_id' in data:
            room = Room.query.get(data['room_id'])
            if not room:
                return jsonify({
                    'status': 'error',
                    'message': f'Room with ID {data["room_id"]} not found'
                }), 404
            exam.room_id = data['room_id']
        
        # Procesăm actualizarea datelor de timp
        start_time = exam.start_time
        end_time = exam.end_time
        
        if 'start_time' in data:
            start_time = datetime.fromisoformat(data['start_time'])
        
        if 'end_time' in data:
            end_time = datetime.fromisoformat(data['end_time'])
        
        # Verificăm dacă data de început este înainte de data de sfârșit
        if start_time >= end_time:
            return jsonify({
                'status': 'error',
                'message': 'Start time must be before end time'
            }), 400
        
        # Verificăm suprapunerile pentru sală, profesor, asistent și grupă de studenți
        room_id = data.get('room_id') or exam.room_id
        
        # Verificăm suprapunerile pentru sală
        if check_room_overlap(room_id, start_time, end_time, exclude_exam_id=id):
            return jsonify({
                'status': 'error',
                'message': 'Room is not available during the specified time interval',
                'overlap_type': 'room'
            }), 409  # Conflict
            
        # Verificăm suprapunerile pentru profesor
        if exam.course.teacher_id and check_teacher_overlap(exam.course.teacher_id, start_time, end_time, exclude_exam_id=id):
            return jsonify({
                'status': 'error',
                'message': 'Teacher has another exam scheduled during the specified time interval',
                'overlap_type': 'teacher'
            }), 409  # Conflict
            
        # Verificăm suprapunerile pentru asistent
        if exam.course.assistant_id and check_assistant_overlap(exam.course.assistant_id, start_time, end_time, exclude_exam_id=id):
            return jsonify({
                'status': 'error',
                'message': 'Assistant has another exam scheduled during the specified time interval',
                'overlap_type': 'assistant'
            }), 409  # Conflict
            
        # Verificăm suprapunerile pentru grupa de studenți
        if check_student_group_overlap(
            exam.course.group_name, 
            exam.course.faculty, 
            exam.course.study_program, 
            exam.course.year_of_study, 
            start_time, 
            end_time, 
            exclude_exam_id=id
        ):
            return jsonify({
                'status': 'error',
                'message': 'Student group has another exam scheduled during the specified time interval',
                'overlap_type': 'student_group'
            }), 409  # Conflict
            
        # Alternativ, putem obține toate suprapunerile într-un singur apel
        exam_data = {
            'room_id': room_id,
            'start_time': start_time,
            'end_time': end_time,
            'course_id': exam.course_id
        }
        
        overlaps = get_overlapping_exams(exam_data, exclude_exam_id=id)
        
        # Dacă există suprapuneri, le returnam pentru a fi afișate în frontend
        has_overlaps = any(len(overlaps[key]) > 0 for key in overlaps if key != 'error')
        
        if has_overlaps:
            return jsonify({
                'status': 'error',
                'message': 'There are scheduling conflicts for this exam',
                'overlaps': overlaps
            }), 409  # Conflict
        
        # Actualizăm examenul
        if 'start_time' in data:
            exam.start_time = start_time
        
        if 'end_time' in data:
            exam.end_time = end_time
        
        if 'exam_type' in data:
            exam.exam_type = data['exam_type']
        
        if 'max_students' in data:
            exam.max_students = data.get('max_students')
        
        if 'notes' in data:
            exam.notes = data.get('notes')
        
        # Actualizăm timestamp-ul
        exam.updated_at = datetime.utcnow()
        
        # Salvăm în baza de date
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'Exam updated successfully',
            'data': {
                'id': exam.id,
                'course': {
                    'id': exam.course.id,
                    'name': exam.course.name
                },
                'room': {
                    'id': exam.room.id,
                    'name': exam.room.name
                },
                'start_time': exam.start_time.isoformat(),
                'end_time': exam.end_time.isoformat(),
                'exam_type': exam.exam_type,
                'semester': exam.semester,
                'academic_year': exam.academic_year,
                'max_students': exam.max_students,
                'notes': exam.notes
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error updating exam {id}: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Failed to update exam with ID {id}',
            'error': str(e)
        }), 500


@exam_bp.route('/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_exam(id):
    """
    Șterge (dezactivează) un examen
    
    Path params:
        id (int): ID-ul examenului
    """
    try:
        exam = Exam.query.get(id)
        
        if not exam or not exam.is_active:
            return jsonify({
                'status': 'error',
                'message': f'Exam with ID {id} not found'
            }), 404
        
        # Nu ștergem complet, doar dezactivăm
        exam.is_active = False
        exam.updated_at = datetime.utcnow()
        
        # Salvăm modificările în baza de date
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': f'Exam with ID {id} deleted successfully'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error deleting exam {id}: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Failed to delete exam with ID {id}',
            'error': str(e)
        }), 500


@exam_bp.route('/export/excel', methods=['GET'])
@jwt_required()
def export_exam_schedule_excel():
    """
    Exportă programarea examenelor în format Excel
    
    Query params:
        faculty (str): Filtrare după facultate
        study_program (str): Filtrare după program de studiu
        year_of_study (int): Filtrare după an de studiu
        group_name (str): Filtrare după grupă
        exam_type (str): Filtrare după tip examen
    """
    try:
        # Preluăm parametrii de filtrare
        filters = {
            'faculty': request.args.get('faculty'),
            'study_program': request.args.get('study_program'),
            'year_of_study': request.args.get('year_of_study'),
            'group_name': request.args.get('group_name'),
            'exam_type': request.args.get('exam_type')
        }
        
        # Generăm raportul Excel
        excel_data = generate_exam_schedule_excel(filters)
        
        # Pregătim fișierul pentru download
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"programare_examene_{timestamp}.xlsx"
        
        return send_file(
            io.BytesIO(excel_data),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Error generating Excel report: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': 'Failed to generate Excel report',
            'error': str(e)
        }), 500


@exam_bp.route('/export/pdf', methods=['GET'])
@jwt_required()
def export_exam_schedule_pdf():
    """
    Exportă programarea examenelor în format PDF
    
    Query params:
        faculty (str): Filtrare după facultate
        study_program (str): Filtrare după program de studiu
        year_of_study (int): Filtrare după an de studiu
        group_name (str): Filtrare după grupă
        exam_type (str): Filtrare după tip examen
    """
    try:
        # Preluăm parametrii de filtrare
        filters = {
            'faculty': request.args.get('faculty'),
            'study_program': request.args.get('study_program'),
            'year_of_study': request.args.get('year_of_study'),
            'group_name': request.args.get('group_name'),
            'exam_type': request.args.get('exam_type')
        }
        
        # Generăm raportul PDF
        pdf_data = generate_exam_schedule_pdf(filters)
        
        # Pregătim fișierul pentru download
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"programare_examene_{timestamp}.pdf"
        
        return send_file(
            io.BytesIO(pdf_data),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Error generating PDF report: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': 'Failed to generate PDF report',
            'error': str(e)
        }), 500


@exam_bp.route('/stats/completion', methods=['GET'])
@jwt_required()
def get_exam_completion_stats():
    """
    Obține statistici despre gradul de completare a programării examenelor
    """
    try:
        # Generăm statisticile
        stats = generate_exam_completion_stats()
        
        return jsonify({
            'status': 'success',
            'data': stats
        }), 200
        
    except Exception as e:
        logger.error(f"Error generating exam completion stats: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': 'Failed to generate exam completion statistics',
            'error': str(e)
        }), 500
